# -*- coding:utf-8 -*-
from dotenv import load_dotenv
load_dotenv()
import os
from woocommerce import API
import json
import io
import re
import unicodedata
import openpyxl
from datetime import datetime

'''
EXPLICAÇÃO - CASO REAL

No momento em que escrevo esse código, presto serviços pra uma determinada empresa que precisa de dados de vendas feitas no
ecommerce e processadas por um determinado gateway de pagamentos.
Acontece que no ecommerce não tem todos os dados necessários, como o código de autorização da vendae a data de previsão
para receber o primeiro valor.

Nesse caso, eu:
- Puxo os dados da venda do nosso site via API do WooCommerce;
- Extraio e transformo os dados de um relatório que exporto em formato '.xls' do gateway de pagamentos porque ele não
possui API disponível;
- Unifico os dados relevantes baseado num identificador que existe tanto no ecommerce quanto no gateway (nsu ou _fixpay_tid);
- Exporto os dados para excel.

OBS: Esse código processa apenas vendas feitas no cartão e recebidas no gateway.
'''

# Informações necessárias para consumir a API do WooCommerce
WOOUSER = os.getenv('WOOUSER')
WOOPASSWORD = os.getenv('WOOPASSWORD')
WOOURL = os.getenv('WOOURL')
WOOVERSION = os.getenv('WOOVERSION')


# EXTRAÇÃO

# Classe de requisição de dados dos pedidos no WooCommerce
class RequestWooOrders:
	def __init__(self, user, password, url, version):
		self.validOrders = []
		self.wcapi = API(
    		url=url,
    		consumer_key=user,
    		consumer_secret=password,
    		version=version
		)

	def request(self, after, before=None):
		
		'''
		#DADOS TESTE
		self.validOrders = [
			{'id': 29373, 'total': '450.00', 'status': 'on-hold', 'billing': {'first_name': 'Santandero', 'last_name': 'Deverus', 'email': 'santanderodeverus@hotmail.com'}, 'meta_data': [{'id': 142379, 'key': 'is_vat_exempt', 'value': 'no'}, {'id': 142380, 'key': '_fixpay_tid', 'value': '10462310191849228184'}, {'id': 142381, 'key': '_fixpay_value', 'value': '450'}]},
			{'id': 29372, 'total': '450.00', 'status': 'on-hold', 'billing': {'first_name': 'Alinelson', 'last_name': 'Logic', 'email': 'magicalin@hotmail.com'}, 'meta_data': [{'id': 142297, 'key': 'is_vat_exempt', 'value': 'no'}, {'id': 142298, 'key': '_fixpay_tid', 'value': '10462310191849228185'}, {'id': 142299, 'key': '_fixpay_value', 'value': '450'}]},
			{'id': 29371, 'total': '450.00', 'status': 'on-hold', 'billing': {'first_name': 'Lovely', 'last_name': 'Luz', 'email': 'lovelyluz@yahoo.com.br'}, 'meta_data': [{'id': 142249, 'key': 'is_vat_exempt', 'value': 'no'}, {'id': 142250, 'key': '_fixpay_tid', 'value': '10462310191849228186'}, {'id': 142251, 'key': '_fixpay_value', 'value': '450'}]}
		]
		return self.validOrders
		'''

		url = 'orders'

		parameters = {
			"after":after,
			"per_page":10,
		}

		if before:
			parameters['before'] = before

		i = 1
		while True:
			parameters["page"] = i

			response = self.wcapi.get(url, params=parameters).json()
			
			if response != []:
				self.validOrders.extend(response)
			else:
				break

			i+=1
		return self.validOrders


# Função de extração de dados de relatórios de venda da FixPay
def fixPayExtract(xl):
	def normalize_str(inputStr, removeSpaces=True, lower=True):
		newStr = ''.join(ch for ch in unicodedata.normalize('NFKD', inputStr) if not unicodedata.combining(ch))
		newStr = newStr.strip()
		if removeSpaces:
			newStr = newStr.replace(' ', '_')
		
		if lower:
			newStr = newStr.lower()
		return newStr
	
	with io.open(xl, 'r', encoding='utf-8') as f:
		readed = f.read()

		readed = (readed.split('<table>')[1]).split('</table>')[0]

		#Header
		header = (readed.split('<thead>')[1]).split('</thead>')[0]
		header = (header.split('<tr>')[1]).split('</tr>')[0]
		header = header.split('</th><th>')
		header = [header[0].replace('<th>','')] + header[1:-1] + [header[-1].replace('</th>','')]
		header = [normalize_str(h) for h in header]
		
		#Body
		body = (readed.split('<tbody>')[1]).split('</tbody>')[0]
		body = body.split('</tr><tr>')
		body = [b.split('</td><td>') for b in body]
		body = [[re.sub('<td>|</td>|<tr>|</tr>', '', i) for i in b] for b in body]		
		
		if len(header) != len(body[0]):
			return False

		dados = [dict([tuple([header[i], e]) for i, e in enumerate(b)]) for b in body]

		return dados

# TRANSFORMAÇÃO

# Selecionando informações relevantes de pedidos
def transformOrders(noProcessedOrders):
	def getMetaData(order, key):
		for meta_data in order['meta_data']:
			if meta_data['key'] == key:
				return meta_data['value']
		return None

	orders = []
	for order in noProcessedOrders:
		currentOrder = {}

		for col in ['id', 'total', 'status']:
			currentOrder[col] = order[col]
		
		for col in ['first_name', 'last_name', 'email']:
			currentOrder[col] = order['billing'][col]

		currentOrder['_fixpay_tid'] = getMetaData(order, '_fixpay_tid')
		
		orders.append(currentOrder)
	return orders

# Unificando pedidos e pagamentos
def joinOrdersAndPayments(orders, dataCred):
	ordersWithPayment = []

	for order in orders:
		for payment in dataCred:
			payment['nsu'] = ''.join([c for c in payment['nsu'] if c.isdigit()])
			if (payment['nsu'] == order['_fixpay_tid']) and (order['_fixpay_tid'] != None):
				currentOrder = order.copy()				
				currentOrder.update(dict([tuple([col, payment[col]]) for col in ['natura', 'bandeira', 'data_da_venda', 'autorizacao', 'data_previsao_cliente']]))
				ordersWithPayment.append(currentOrder)
				break
	return ordersWithPayment

# Excrevendo dados em excel
def writeInXL(ordersWithPayment):
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Pedidos no Cartão"

	for i, key in enumerate(ordersWithPayment[0].keys()):
		ws.cell(row=1, column=i+1, value=key)

	for i, order in enumerate(ordersWithPayment):
		for j, value in enumerate(order.values()):
			ws.cell(row=i+2, column=j+1, value=value)
	
	dt_string = datetime.now().strftime("%d_%m_%Y-%H_%M_%S")
	
	wb.save(f"vendas_{dt_string}.xlsx")

def main():

	#Requisitando dados do ecommerce
	requestWooOrders = RequestWooOrders(WOOUSER, WOOPASSWORD, WOOURL, WOOVERSION)
	requestWooOrders.request(after="2023-10-01T00:00:00")

	#Extraindo dados do relatório do gateway
	dataCred = fixPayExtract("cred.xls")

	#Transformando dados de pedidos
	orders = transformOrders(requestWooOrders.validOrders)

	#Unificando dados de pedidos e pagamentos
	ordersWithPayment = joinOrdersAndPayments(orders, dataCred)

	#Escrevendo dados em excel
	writeInXL(ordersWithPayment)
main()